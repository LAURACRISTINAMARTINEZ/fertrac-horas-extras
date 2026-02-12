import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, date, time, timedelta
import matplotlib.pyplot as plt
from io import BytesIO
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import socket  # Para manejar errores de conexión

st.set_page_config(
    page_title="Calculadora Horas Extras Fertrac",
    layout="wide",
    page_icon="🕒"
)

# ============================================================================
# CONFIGURACIÓN DE ENVÍO DE CORREO
# ============================================================================
def enviar_notificacion_email(destinatario, asunto, cuerpo_html, archivo_adjunto=None, nombre_archivo=None):
    """
    Envía un correo de notificación cada vez que se usa la herramienta.
    Las credenciales SMTP se configuran en la barra lateral o en st.secrets.
    """
    try:
        # Obtener credenciales SMTP desde session_state (barra lateral)
        smtp_server = st.session_state.get("smtp_server", "smtp.gmail.com")
        smtp_port = int(st.session_state.get("smtp_port", 587))
        smtp_user = st.session_state.get("smtp_user", "analista_automatizacion@fertrac.com")
        smtp_password = st.session_state.get("smtp_password", "ssrz ldin nvyx ixry")
        
        # También intentar leer desde secrets (para despliegue en producción)
        try:
            smtp_server = st.secrets.get("SMTP_SERVER", smtp_server)
            smtp_port = int(st.secrets.get("SMTP_PORT", smtp_port))
            smtp_user = st.secrets.get("SMTP_USER", smtp_user)
            smtp_password = st.secrets.get("SMTP_PASSWORD", smtp_password)
        except Exception:
            pass  # Si no hay archivo secrets.toml, se usan los valores del sidebar
        
        if not smtp_user or not smtp_password:
            return False, "Credenciales SMTP no configuradas"
        
        # Crear el mensaje de correo
        msg = MIMEMultipart("alternative")
        msg["From"] = smtp_user
        msg["To"] = destinatario
        msg["Subject"] = asunto
        
        # Agregar el cuerpo del correo en formato HTML
        msg.attach(MIMEText(cuerpo_html, "html"))
        
        # Adjuntar archivo si se proporcionó (opcional)
        if archivo_adjunto and nombre_archivo:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(archivo_adjunto)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={nombre_archivo}")
            msg.attach(part)
        
        # Conectar al servidor SMTP y enviar el correo
        with smtplib.SMTP(smtp_server, smtp_port, timeout=10) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_user, destinatario, msg.as_string())
        
        return True, "Correo enviado exitosamente"
    
    except smtplib.SMTPAuthenticationError:
        return False, "Error de autenticación SMTP. Verifica usuario y contraseña."
    except socket.timeout:
        return False, "Tiempo de espera agotado al conectar con el servidor SMTP."
    except Exception as e:
        return False, f"Error al enviar correo: {str(e)}"

def construir_cuerpo_email(num_registros, num_empleados, areas, fecha_ejecucion, total_valor_extras):
    """Construye el cuerpo HTML del correo de notificación con el resumen del cálculo."""
    return f"""
    <html>
    <body style="font-family: Arial, sans-serif; color: #333;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
            <div style="background-color: #f37021; padding: 15px; border-radius: 8px 8px 0 0;">
                <h2 style="color: white; margin: 0;">🕒 Notificación - Calculadora Horas Extras Fertrac</h2>
            </div>
            <div style="border: 1px solid #ddd; border-top: none; padding: 20px; border-radius: 0 0 8px 8px;">
                <p>Se ha realizado un nuevo cálculo de horas extras con los siguientes detalles:</p>
                
                <table style="width: 100%; border-collapse: collapse; margin: 15px 0;">
                    <tr style="background-color: #f8f8f8;">
                        <td style="padding: 10px; border: 1px solid #ddd; font-weight: bold;">Fecha y hora de ejecución</td>
                        <td style="padding: 10px; border: 1px solid #ddd;">{fecha_ejecucion}</td>
                    </tr>
                    <tr>
                        <td style="padding: 10px; border: 1px solid #ddd; font-weight: bold;">Registros procesados</td>
                        <td style="padding: 10px; border: 1px solid #ddd;">{num_registros}</td>
                    </tr>
                    <tr style="background-color: #f8f8f8;">
                        <td style="padding: 10px; border: 1px solid #ddd; font-weight: bold;">Empleados únicos</td>
                        <td style="padding: 10px; border: 1px solid #ddd;">{num_empleados}</td>
                    </tr>
                    <tr>
                        <td style="padding: 10px; border: 1px solid #ddd; font-weight: bold;">Áreas involucradas</td>
                        <td style="padding: 10px; border: 1px solid #ddd;">{areas}</td>
                    </tr>
                    <tr style="background-color: #f8f8f8;">
                        <td style="padding: 10px; border: 1px solid #ddd; font-weight: bold;">Valor total extras</td>
                        <td style="padding: 10px; border: 1px solid #ddd; color: #f37021; font-weight: bold;">${total_valor_extras:,.2f}</td>
                    </tr>
                </table>
                
                <p style="color: #888; font-size: 12px; margin-top: 20px;">
                    Este es un correo automático generado por la Calculadora de Horas Extras de Fertrac.
                </p>
            </div>
        </div>
    </body>
    </html>
    """

# Inicializar estado para evitar reenvíos de correo en la misma sesión
if "email_enviado" not in st.session_state:
    st.session_state.email_enviado = False

# ============================================================================
# CONFIGURACIÓN SMTP EN LA BARRA LATERAL
# ============================================================================
with st.sidebar:
    st.markdown("### ⚙️ Configuración de correo")
    st.caption("Configura las credenciales SMTP para recibir notificaciones por correo.")
    
    with st.expander("🔧 Credenciales SMTP", expanded=False):
        st.session_state["smtp_server"] = st.text_input(
            "Servidor SMTP", 
            value=st.session_state.get("smtp_server", "smtp.gmail.com"),
            help="Ej: smtp.gmail.com, smtp.office365.com"
        )
        st.session_state["smtp_port"] = st.text_input(
            "Puerto SMTP", 
            value=st.session_state.get("smtp_port", "587"),
            help="Normalmente 587 para TLS"
        )
        st.session_state["smtp_user"] = st.text_input(
            "Correo remitente", 
            value=st.session_state.get("smtp_user", ""),
            help="El correo desde el que se enviarán las notificaciones"
        )
        st.session_state["smtp_password"] = st.text_input(
            "Contraseña / App Password", 
            type="password",
            value=st.session_state.get("smtp_password", ""),
            help="Para Gmail usa una 'App Password' (contraseña de aplicación)"
        )
        
        st.info("""
        💡 **Para Gmail:** Usa una contraseña de aplicación.
        Ve a myaccount.google.com → Seguridad → Contraseñas de aplicaciones.
        
        **Para Outlook/Office365:** smtp.office365.com, puerto 587.
        """)
    
    envio_habilitado = bool(st.session_state.get("smtp_user")) and bool(st.session_state.get("smtp_password"))
    if envio_habilitado:
        st.success("✅ Credenciales SMTP configuradas")
    else:
        st.warning("⚠️ Configura las credenciales SMTP para habilitar notificaciones por correo")

st.image("logo_fertrac.png", width=200)
st.markdown(
    "<h2 style='color:#f37021;'>Bienvenido a la herramienta de cálculo de horas extras de Fertrac.</h2>"
    "<p>Por favor, sube los archivos requeridos para comenzar.</p>",
    unsafe_allow_html=True
)

col1, col2 = st.columns(2)

with col1:
    input_file = st.file_uploader("📤 Subir archivo de datos (input_datos.xlsx)", type=["xlsx"], key="input")
    empleados_file = st.file_uploader("📤 Subir base de empleados", type=["xlsx"], key="empleados")

with col2:
    porcentaje_file = st.file_uploader("📤 Subir factores de horas extras (%)", type=["xlsx"], key="porcentaje")
    turnos_file = st.file_uploader("📤 Subir configuración de turnos", type=["xlsx"], key="turnos")

@st.cache_data
def cargar_excel(file):
    """Cachea la lectura de archivos Excel para no re-leer en cada interacción"""
    return pd.read_excel(file)

if input_file and empleados_file and porcentaje_file and turnos_file:
    df_input = cargar_excel(input_file)
    df_empleados = cargar_excel(empleados_file)
    df_porcentaje = cargar_excel(porcentaje_file)
    df_turnos = cargar_excel(turnos_file)

    # Normalizar columnas
    df_input.columns = df_input.columns.str.upper().str.strip()
    df_empleados.columns = df_empleados.columns.str.upper().str.strip()
    df_porcentaje.columns = df_porcentaje.columns.str.upper().str.strip()
    df_turnos.columns = df_turnos.columns.str.upper().str.strip()

    # Normalizar nombres de columnas de cédula
    # El archivo de input puede tener "CÉDULA" y el de empleados "CEDULA"
    if "CÉDULA" in df_input.columns:
        cedula_input = "CÉDULA"
    elif "CEDULA" in df_input.columns:
        cedula_input = "CEDULA"
    else:
        st.error("⚠️ Error: No se encontró columna de cédula en archivo de datos")
        st.info("La columna debe llamarse 'CÉDULA' o 'CEDULA'")
        st.stop()
    
    if "CEDULA" in df_empleados.columns:
        cedula_empleados = "CEDULA"
    elif "CÉDULA" in df_empleados.columns:
        cedula_empleados = "CÉDULA"
    else:
        st.error("⚠️ Error: No se encontró columna de cédula en base de empleados")
        st.info("La columna debe llamarse 'CEDULA' o 'CÉDULA'")
        st.stop()
    
    # Convertir cédulas a string para evitar problemas de tipo
    df_input[cedula_input] = df_input[cedula_input].astype(str).str.strip()
    df_empleados[cedula_empleados] = df_empleados[cedula_empleados].astype(str).str.strip()
    
    # Merge con empleados
    df = df_input.merge(df_empleados, left_on=cedula_input, right_on=cedula_empleados, how="left")
    
    # Validar que existan las columnas necesarias del archivo de empleados
    columnas_requeridas = ["NOMBRE", "AREA", "SALARIO BASICO"]
    columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
    
    if columnas_faltantes:
        st.error(f"⚠️ Error: Faltan columnas en el archivo de empleados: {', '.join(columnas_faltantes)}")
        st.info(f"""
        El archivo de empleados debe tener las siguientes columnas:
        - CEDULA o CÉDULA
        - NOMBRE
        - AREA
        - SALARIO BASICO
        - CARGO (opcional, pero recomendado)
        """)
        st.stop()
    
    # Verificar si existe CARGO (opcional pero recomendado)
    if "CARGO" not in df.columns:
        st.warning("⚠️ Advertencia: No se encontró columna 'CARGO' en el archivo de empleados. Se dejará vacía.")
        df["CARGO"] = ""
    
    # COMISIÓN O BONIFICACIÓN debe venir del archivo input_datos
    # Si no existe en input_datos, llenar con 0
    if "COMISIÓN O BONIFICACIÓN" not in df.columns:
        if "COMISION O BONIFICACION" not in df.columns:
            st.warning("⚠️ Advertencia: No se encontró columna 'COMISIÓN O BONIFICACIÓN' en el archivo de datos de entrada. Se asumirá valor $0 para todos.")
            df["COMISIÓN O BONIFICACIÓN"] = 0
        else:
            df["COMISIÓN O BONIFICACIÓN"] = df["COMISION O BONIFICACION"]
    
    # Agregar columna OBSERVACIONES si no existe en el input
    if "OBSERVACIONES" not in df.columns and "OBSERVACION" not in df.columns:
        df["OBSERVACIONES"] = ""  # Columna vacía por defecto
    elif "OBSERVACION" in df.columns:
        df["OBSERVACIONES"] = df["OBSERVACION"]  # Renombrar si existe con otro nombre
    
    # Verificar que se hayan encontrado coincidencias
    # Primero verificar si existe la columna NOMBRE después del merge
    if "NOMBRE" in df.columns:
        empleados_sin_info = df[df["NOMBRE"].isna()]
        if len(empleados_sin_info) > 0:
            st.warning(f"⚠️ Advertencia: {len(empleados_sin_info)} registros no tienen información de empleado")
            st.warning(f"Cédulas sin coincidencia: {empleados_sin_info[cedula_input].unique().tolist()[:5]}")
            st.info("Verifica que las cédulas en ambos archivos coincidan exactamente")
    else:
        st.error("⚠️ Error: No se encontró la columna NOMBRE en el archivo de empleados")
        st.info("El archivo de empleados debe tener una columna llamada 'NOMBRE' o 'Nombre'")
        st.stop()

    # Procesamiento de fechas y horas
    try:
        df["FECHA"] = pd.to_datetime(df["FECHA"], errors='coerce')
        
        # Verificar que las fechas se convirtieron correctamente
        if df["FECHA"].isna().any():
            fechas_problema = df[df["FECHA"].isna()].index.tolist()
            st.error(f"⚠️ Error: Algunas fechas no tienen el formato correcto en las filas: {fechas_problema[:5]}")
            st.info("Las fechas deben estar en formato: YYYY-MM-DD (ej: 2025-02-04) o DD/MM/YYYY")
            st.stop()
        
        df["DIA_NUM"] = df["FECHA"].dt.weekday  # 0=Lun, 5=Sáb, 6=Dom
        
        # Crear columna DÍA con nombre del día en español
        dias_espanol = {
            0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves',
            4: 'Viernes', 5: 'Sábado', 6: 'Domingo'
        }
        df["DÍA"] = df["DIA_NUM"].map(dias_espanol)
    except Exception as e:
        st.error(f"⚠️ Error al procesar fechas: {str(e)}")
        st.info("Verifica que la columna FECHA tenga fechas válidas")
        st.stop()

    try:
        # Función para convertir horas de cualquier formato a time
        def convertir_a_time(valor):
            if pd.isna(valor):
                return None
            # Si ya es un objeto time, devolverlo
            if isinstance(valor, time):
                return valor
            # Si es datetime, extraer solo la hora
            if isinstance(valor, datetime):
                return valor.time()
            # Si es string, intentar parsearlo
            if isinstance(valor, str):
                try:
                    # Intentar parsear como HH:MM o HH:MM:SS
                    parsed = pd.to_datetime(valor, format='%H:%M', errors='coerce')
                    if pd.isna(parsed):
                        parsed = pd.to_datetime(valor, format='%H:%M:%SS', errors='coerce')
                    if not pd.isna(parsed):
                        return parsed.time()
                except:
                    pass
            return None
        
        # Convertir horas usando la función robusta
        df["HRA INGRESO"] = df["HRA INGRESO"].apply(convertir_a_time)
        df["HORA SALIDA"] = df["HORA SALIDA"].apply(convertir_a_time)
        
        # Verificar que las conversiones fueron exitosas
        if df["HRA INGRESO"].isna().any() or df["HORA SALIDA"].isna().any():
            st.error("⚠️ Error: Algunas horas de entrada o salida no tienen el formato correcto")
            st.info("Las horas pueden estar en formato Excel (tiempo) o texto HH:MM (ej: 08:00, 14:00, 18:30)")
            st.stop()
    except Exception as e:
        st.error(f"⚠️ Error al procesar horas: {str(e)}")
        st.info("Verifica el formato de las columnas HRA INGRESO y HORA SALIDA")
        st.stop()

    def convertir_a_datetime(fecha, hora):
        return pd.to_datetime(fecha.astype(str) + ' ' + hora.astype(str))

    df["DT_INGRESO"] = convertir_a_datetime(df["FECHA"], df["HRA INGRESO"])
    df["DT_SALIDA"] = convertir_a_datetime(df["FECHA"], df["HORA SALIDA"])

    # Crear diccionario de configuración de turnos
    turnos_config = {}
    for _, row in df_turnos.iterrows():
        turno_nombre = str(row["TURNO"]).upper().strip()
        
        # Convertir horas de manera segura (acepta time, datetime, o string)
        try:
            # Función para convertir a time
            def a_time(valor):
                if isinstance(valor, time):
                    return valor
                if isinstance(valor, datetime):
                    return valor.time()
                if isinstance(valor, str):
                    parsed = pd.to_datetime(valor, format='%H:%M', errors='coerce')
                    if pd.isna(parsed):
                        parsed = pd.to_datetime(valor, format='%H:%M:%S', errors='coerce')
                    if not pd.isna(parsed):
                        return parsed.time()
                return None
            
            hora_entrada = a_time(row["HORA ENTRADA"])
            hora_salida = a_time(row["HORA SALIDA"])
            
            if hora_entrada is None or hora_salida is None:
                st.warning(f"⚠️ Advertencia: Turno '{turno_nombre}' tiene formato de hora inválido, se omitirá")
                continue
            
            turnos_config[turno_nombre] = {
                "entrada": hora_entrada,
                "salida": hora_salida
            }
        except Exception as e:
            st.error(f"Error procesando turno {turno_nombre}: {e}")
            continue
    
    # Verificar que se cargaron turnos
    if not turnos_config:
        st.error("⚠️ No se pudo cargar ningún turno del archivo de configuración")
        st.stop()
    
    # Mostrar configuración de turnos cargada
    st.success(f"""
    ✅ **Configuración de turnos cargada:**
    {chr(10).join([f"- {turno}: {config['entrada'].strftime('%H:%M')} a {config['salida'].strftime('%H:%M')}" 
                   for turno, config in turnos_config.items()])}
    """)

    # Definir horarios según turno (usando configuración)
    def obtener_horarios_turno(row):
        """Retorna las horas de entrada y salida según la configuración del turno"""
        turno = row.get("TURNO", "TURNO 1").upper().strip()
        
        # Si el turno existe en la configuración, usar esos horarios
        if turno in turnos_config:
            hora_entrada = datetime.combine(row["FECHA"], turnos_config[turno]["entrada"])
            hora_salida = datetime.combine(row["FECHA"], turnos_config[turno]["salida"])
        else:
            # Si no está configurado, usar valores por defecto de TURNO 1
            st.warning(f"⚠️ Turno '{turno}' no encontrado en configuración. Usando TURNO 1 por defecto.")
            hora_entrada = datetime.combine(row["FECHA"], time(8, 0))
            hora_salida = datetime.combine(row["FECHA"], time(18, 0))
        
        return hora_entrada, hora_salida

    # ============================================================================
    # FUNCIÓN CORREGIDA - Comparación de horas sin confundir fechas
    # ============================================================================
    def es_horario_nocturno(hora):
        """
        Determina si una hora (objeto time) está en horario nocturno.
        Horario nocturno: 19:00 (7 PM) a 06:00 (6 AM)
        
        CORRECCIÓN: Compara solo las HORAS, no fechas completas.
        """
        # Extraer solo la hora si es datetime
        if isinstance(hora, datetime):
            hora = hora.time()
        
        # Nocturno: >= 19:00 (7 PM en adelante) O < 06:00 (antes de 6 AM)
        return hora >= time(19, 0) or hora < time(6, 0)

    def calcular_horas_extras_y_recargo(row):
        """
        Calcula horas extras diurnas, nocturnas y recargo nocturno de forma precisa.
        
        DEFINICIONES:
        - Horario NOCTURNO: 19:00 (7 PM) hasta 06:00 (6 AM)
        - Horario DIURNO: 06:00 (6 AM) hasta 19:00 (7 PM)
        
        CATEGORÍAS:
        1. RECARGO NOCTURNO: Horas DENTRO de la jornada normal trabajadas en horario nocturno
        2. EXTRA DIURNA: Horas FUERA de la jornada normal trabajadas en horario diurno
        3. EXTRA NOCTURNA: Horas FUERA de la jornada normal trabajadas en horario nocturno
        
        CASO ESPECIAL: Cuando el trabajador tiene una franja horaria de madrugada
        (ej: 00:01 a 03:00), significa que trabajó después de medianoche como continuación
        del día anterior. En este caso, TODO el rango es hora extra nocturna.
        """
        ingreso_real = row["DT_INGRESO"]
        salida_real = row["DT_SALIDA"]
        
        # Obtener horarios del turno esperado
        hora_entrada_esperada, hora_salida_esperada = obtener_horarios_turno(row)
        
        # Inicializar contadores
        horas_extra_diurna = 0
        horas_extra_nocturna = 0
        recargo_nocturno = 0
        
        # ========================================================================
        # CASO ESPECIAL: Franja completamente en madrugada (cruce de medianoche)
        # Si la hora de ingreso Y la hora de salida son ambas antes de la hora
        # de entrada del turno, y están en horario nocturno, entonces TODO
        # el rango es hora extra nocturna (el trabajador pasó de un día al otro).
        # ========================================================================
        hora_ingreso = ingreso_real.time()
        hora_salida = salida_real.time()
        hora_entrada_turno = hora_entrada_esperada.time()
        
        # Detectar: ambas horas antes del inicio del turno Y en horario nocturno (madrugada)
        if (hora_ingreso < hora_entrada_turno and hora_salida <= hora_entrada_turno
            and hora_salida > hora_ingreso
            and es_horario_nocturno(hora_ingreso) and es_horario_nocturno(hora_salida)):
            # Todo es extra nocturna pura
            horas_extra_nocturna = (salida_real - ingreso_real).total_seconds() / 3600
            return max(horas_extra_diurna, 0), max(horas_extra_nocturna, 0), max(recargo_nocturno, 0)
        
        # ========================================================================
        # PARTE 1: RECARGO NOCTURNO (horas normales en horario nocturno)
        # ========================================================================
        # El recargo se aplica a las horas DENTRO de la jornada que caen en horario nocturno
        
        # Determinar el rango de jornada normal (lo que realmente trabajó dentro del turno)
        inicio_jornada_normal = max(ingreso_real, hora_entrada_esperada)
        fin_jornada_normal = min(salida_real, hora_salida_esperada)
        
        if inicio_jornada_normal < fin_jornada_normal:
            # Hay jornada normal trabajada, calcular cuánto de ella es nocturno
            tiempo_actual = inicio_jornada_normal
            
            while tiempo_actual < fin_jornada_normal:
                hora_actual = tiempo_actual.time()
                
                if es_horario_nocturno(hora_actual):
                    # Esta hora está en horario nocturno
                    if hora_actual >= time(19, 0):
                        # Entre 19:00 y 23:59:59
                        fin_segmento = datetime.combine(tiempo_actual.date(), time(23, 59, 59))
                    else:
                        # Entre 00:00 y 05:59:59
                        fin_segmento = datetime.combine(tiempo_actual.date(), time(6, 0))
                    
                    fin_segmento = min(fin_segmento, fin_jornada_normal)
                    horas = (fin_segmento - tiempo_actual).total_seconds() / 3600
                    recargo_nocturno += horas
                    tiempo_actual = fin_segmento
                    
                    # Si llegamos al final del día, avanzar a medianoche
                    if fin_segmento.time() == time(23, 59, 59) and tiempo_actual < fin_jornada_normal:
                        tiempo_actual = tiempo_actual + timedelta(seconds=1)
                else:
                    # Esta hora es diurna, avanzar hasta el inicio del horario nocturno
                    fin_segmento = datetime.combine(tiempo_actual.date(), time(19, 0))
                    fin_segmento = min(fin_segmento, fin_jornada_normal)
                    tiempo_actual = fin_segmento
                
                if tiempo_actual >= fin_jornada_normal:
                    break
        
        # ========================================================================
        # PARTE 2: HORAS EXTRA ANTES DEL TURNO
        # ========================================================================
        if ingreso_real < hora_entrada_esperada:
            tiempo_actual = ingreso_real
            
            while tiempo_actual < hora_entrada_esperada:
                hora_actual = tiempo_actual.time()
                
                if es_horario_nocturno(hora_actual):
                    # Hora extra nocturna
                    if hora_actual >= time(19, 0):
                        # Entre 19:00 y 23:59:59
                        fin_segmento = datetime.combine(tiempo_actual.date(), time(23, 59, 59))
                    else:
                        # Entre 00:00 y 05:59:59
                        fin_segmento = datetime.combine(tiempo_actual.date(), time(6, 0))
                    
                    fin_segmento = min(fin_segmento, hora_entrada_esperada)
                    horas = (fin_segmento - tiempo_actual).total_seconds() / 3600
                    horas_extra_nocturna += horas
                    tiempo_actual = fin_segmento
                    
                    # Si llegamos al final del día, avanzar a medianoche
                    if fin_segmento.time() == time(23, 59, 59) and tiempo_actual < hora_entrada_esperada:
                        tiempo_actual = tiempo_actual + timedelta(seconds=1)
                else:
                    # Hora extra diurna
                    fin_segmento = datetime.combine(tiempo_actual.date(), time(19, 0))
                    fin_segmento = min(fin_segmento, hora_entrada_esperada)
                    horas = (fin_segmento - tiempo_actual).total_seconds() / 3600
                    horas_extra_diurna += horas
                    tiempo_actual = fin_segmento
                
                if tiempo_actual >= hora_entrada_esperada:
                    break
        
        # ========================================================================
        # PARTE 3: HORAS EXTRA DESPUÉS DEL TURNO
        # ========================================================================
        if salida_real > hora_salida_esperada:
            tiempo_actual = hora_salida_esperada
            
            while tiempo_actual < salida_real:
                hora_actual = tiempo_actual.time()
                
                if es_horario_nocturno(hora_actual):
                    # Hora extra nocturna
                    if hora_actual >= time(19, 0):
                        # Entre 19:00 y 23:59:59
                        fin_segmento = datetime.combine(tiempo_actual.date(), time(23, 59, 59))
                    else:
                        # Entre 00:00 y 05:59:59
                        fin_segmento = datetime.combine(tiempo_actual.date(), time(6, 0))
                    
                    fin_segmento = min(fin_segmento, salida_real)
                    horas = (fin_segmento - tiempo_actual).total_seconds() / 3600
                    horas_extra_nocturna += horas
                    tiempo_actual = fin_segmento
                    
                    # Si llegamos al final del día, avanzar a medianoche
                    if fin_segmento.time() == time(23, 59, 59) and tiempo_actual < salida_real:
                        tiempo_actual = tiempo_actual + timedelta(seconds=1)
                else:
                    # Hora extra diurna
                    fin_segmento = datetime.combine(tiempo_actual.date(), time(19, 0))
                    fin_segmento = min(fin_segmento, salida_real)
                    horas = (fin_segmento - tiempo_actual).total_seconds() / 3600
                    horas_extra_diurna += horas
                    tiempo_actual = fin_segmento
                
                if tiempo_actual >= salida_real:
                    break
        
        # Asegurar que no haya valores negativos
        return max(horas_extra_diurna, 0), max(horas_extra_nocturna, 0), max(recargo_nocturno, 0)

    # Agregar columnas con los horarios del turno para visualización
    def obtener_horarios_turno_para_mostrar(row):
        """Retorna los horarios del turno como strings para mostrar en tabla"""
        turno = row.get("TURNO", "TURNO 1").upper().strip()
        
        if turno in turnos_config:
            hora_entrada = turnos_config[turno]["entrada"]
            hora_salida = turnos_config[turno]["salida"]
        else:
            hora_entrada = time(8, 0)
            hora_salida = time(18, 0)
        
        return hora_entrada.strftime('%H:%M'), hora_salida.strftime('%H:%M')
    
    # Aplicar para obtener horarios del turno
    horarios_turno = df.apply(obtener_horarios_turno_para_mostrar, axis=1)
    df["TURNO ENTRADA"] = [h[0] for h in horarios_turno]
    df["TURNO SALIDA"] = [h[1] for h in horarios_turno]
    
    # Aplicar cálculos - VECTORIZADO (sin .apply)
    total_horas = (df["DT_SALIDA"] - df["DT_INGRESO"]).dt.total_seconds() / 3600
    
    # Detectar franjas de madrugada: ingreso y salida antes de las 06:00 AM
    # En estos casos NO se descuenta almuerzo
    hora_ingreso = df["HRA INGRESO"].apply(lambda h: h.hour if h else 0)
    hora_salida_val = df["HORA SALIDA"].apply(lambda h: h.hour if h else 0)
    es_franja_madrugada = (hora_ingreso < 6) & (hora_salida_val <= 6)
    
    # Descontar 1 hora de almuerzo solo de lunes a viernes Y cuando NO es franja de madrugada
    df["HORAS TRABAJADAS"] = np.where(
        (df["DIA_NUM"] < 5) & (~es_franja_madrugada),
        total_horas - 1,
        total_horas
    )
    
    # Calcular horas extras diurnas, nocturnas y recargo nocturno
    resultados = df.apply(calcular_horas_extras_y_recargo, axis=1)
    df["HORAS EXTRA DIURNA"] = [r[0] for r in resultados]
    df["HORAS EXTRA NOCTURNA"] = [r[1] for r in resultados]
    df["RECARGO NOCTURNO"] = [r[2] for r in resultados]
    
    # Total de horas extra (diurnas + nocturnas)
    df["TOTAL HORAS EXTRA"] = df["HORAS EXTRA DIURNA"] + df["HORAS EXTRA NOCTURNA"]

    # Asignar tipos de hora extra
    df["TIPO EXTRA"] = df.apply(
        lambda row: "Recargo Nocturno" if row["RECARGO NOCTURNO"] > 0 else 
                    ("Extra Nocturna" if row["HORAS EXTRA NOCTURNA"] > 0 else "Extra Diurna"),
        axis=1
    )

    # ============================================================================
    # CORRECCIÓN PRINCIPAL: Calcular TOTAL BASE LIQUIDACION PRIMERO
    # ============================================================================
    # TOTAL BASE LIQUIDACION = SALARIO BASICO + COMISION O BONIFICACION
    df["TOTAL BASE LIQUIDACION"] = df["SALARIO BASICO"] + df["COMISIÓN O BONIFICACIÓN"]
    
    # Mapear factores
    factor_map = dict(zip(df_porcentaje["TIPO HORA EXTRA"].str.upper(), df_porcentaje["FACTOR"]))
    
    # Calcular valores monetarios
    # IMPORTANTE: Dividir por 220 horas (aplicable desde 2do semestre 2025)
    # USAR TOTAL BASE LIQUIDACION en lugar de SALARIO BASICO
    df["IMPORTE HORA"] = df["TOTAL BASE LIQUIDACION"] / 220
    
    # Calcular valor de cada tipo de hora
    df["VALOR EXTRA DIURNA"] = df.apply(
        lambda row: row["HORAS EXTRA DIURNA"] * row["IMPORTE HORA"] * factor_map.get("EXTRA DIURNA", 1.25),
        axis=1
    )
    
    df["VALOR EXTRA NOCTURNA"] = df.apply(
        lambda row: row["HORAS EXTRA NOCTURNA"] * row["IMPORTE HORA"] * factor_map.get("EXTRA NOCTURNA", 1.25),
        axis=1
    )
    
    df["VALOR RECARGO NOCTURNO"] = df.apply(
        lambda row: row["RECARGO NOCTURNO"] * row["IMPORTE HORA"] * factor_map.get("RECARGO NOCTURNO", 1.35),
        axis=1
    )
    
    # Valor total de extras (suma de diurna + nocturna + recargo)
    df["VALOR TOTAL EXTRAS"] = df["VALOR EXTRA DIURNA"] + df["VALOR EXTRA NOCTURNA"] + df["VALOR RECARGO NOCTURNO"]

    # Redondear valores MONETARIOS (los cálculos ya se hicieron con precisión completa)
    df["IMPORTE HORA"] = df["IMPORTE HORA"].round(2)
    df["VALOR EXTRA DIURNA"] = df["VALOR EXTRA DIURNA"].round(2)
    df["VALOR EXTRA NOCTURNA"] = df["VALOR EXTRA NOCTURNA"].round(2)
    df["VALOR RECARGO NOCTURNO"] = df["VALOR RECARGO NOCTURNO"].round(2)
    df["VALOR TOTAL EXTRAS"] = df["VALOR TOTAL EXTRAS"].round(2)
    df["TOTAL BASE LIQUIDACION"] = df["TOTAL BASE LIQUIDACION"].round(2)
    
    # Redondear horas SOLO PARA VISUALIZACIÓN (después de los cálculos monetarios)
    # Mantener precisión interna pero mostrar 2 decimales
    df["HORAS TRABAJADAS_DISPLAY"] = df["HORAS TRABAJADAS"].round(2)
    df["HORAS EXTRA DIURNA_DISPLAY"] = df["HORAS EXTRA DIURNA"].round(2)
    df["HORAS EXTRA NOCTURNA_DISPLAY"] = df["HORAS EXTRA NOCTURNA"].round(2)
    df["RECARGO NOCTURNO_DISPLAY"] = df["RECARGO NOCTURNO"].round(2)
    df["TOTAL HORAS EXTRA_DISPLAY"] = df["TOTAL HORAS EXTRA"].round(2)

    # Resumen informativo
    st.info(f"""
    ✅ **Configuración aplicada:**
    - **BASE DE CÁLCULO:** Total Base Liquidación (Salario Básico + Comisión/Bonificación)
    - División por **220 horas** mensuales (vigente desde 2do semestre 2025)
    - Descuento de **1 hora de almuerzo** en días de lunes a viernes
    - **Horario nocturno:** 7:00 PM a 6:00 AM (recargo aplicable a jornada ordinaria en este horario)
    - **Horas extra diurnas:** Fuera de jornada en horario diurno (6:00 AM - 7:00 PM)
    - **Horas extra nocturnas:** Fuera de jornada en horario nocturno (7:00 PM - 6:00 AM)
    - **Turnos configurables** según archivo de turnos cargado
    """)

    # ============================================================================
    # ENVÍO AUTOMÁTICO DE CORREO DE NOTIFICACIÓN
    # ============================================================================
    # Se genera una clave única basada en los datos procesados para evitar
    # que Streamlit reenvíe el correo cada vez que recarga la página
    data_hash = f"{len(df)}_{df[cedula_input].nunique()}_{df['VALOR TOTAL EXTRAS'].sum():.2f}"
    
    if not st.session_state.get(f"email_enviado_{data_hash}", False):
        envio_habilitado = bool(st.session_state.get("smtp_user")) and bool(st.session_state.get("smtp_password"))
        
        if envio_habilitado:
            # Preparar los datos resumen para incluir en el correo
            num_registros = len(df)
            num_empleados = df[cedula_input].nunique()
            areas = ", ".join(df["AREA"].dropna().unique().tolist())
            fecha_ejecucion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            total_valor_extras = df["VALOR TOTAL EXTRAS"].sum()
            
            # Construir el cuerpo HTML del correo
            cuerpo = construir_cuerpo_email(
                num_registros, num_empleados, areas, fecha_ejecucion, total_valor_extras
            )
            
            # Intentar enviar el correo a data_science@fertrac.com
            exito, mensaje = enviar_notificacion_email(
                destinatario="data_science@fertrac.com",
                asunto=f"📊 Cálculo de Horas Extras - {fecha_ejecucion} - {num_empleados} empleados",
                cuerpo_html=cuerpo
            )
            
            # Mostrar resultado al usuario con un toast (notificación pequeña)
            if exito:
                st.toast("📧 Notificación enviada a data_science@fertrac.com", icon="✅")
                st.session_state[f"email_enviado_{data_hash}"] = True
            else:
                st.toast(f"⚠️ No se pudo enviar el correo: {mensaje}", icon="⚠️")

    # Agregar columna de mes y año para análisis temporal
    df["MES"] = df["FECHA"].dt.to_period('M')
    
    # Crear nombres de mes en ESPAÑOL
    meses_espanol = {
        'January': 'Enero',
        'February': 'Febrero',
        'March': 'Marzo',
        'April': 'Abril',
        'May': 'Mayo',
        'June': 'Junio',
        'July': 'Julio',
        'August': 'Agosto',
        'September': 'Septiembre',
        'October': 'Octubre',
        'November': 'Noviembre',
        'December': 'Diciembre'
    }
    
    # Primero crear en inglés y luego traducir
    df["MES_NOMBRE_TEMP"] = df["FECHA"].dt.strftime('%B')
    
    # Traducir a español
    def traducir_mes(mes_nombre):
        for ingles, espanol in meses_espanol.items():
            if ingles in mes_nombre:
                return mes_nombre.replace(ingles, espanol)
        return mes_nombre
    
    df["MES_NOMBRE"] = df["MES_NOMBRE_TEMP"].apply(traducir_mes)
    df = df.drop("MES_NOMBRE_TEMP", axis=1)

    # FILTRO POR ÁREA
    st.subheader("🔍 Filtros de visualización")
    col_filtro1, col_filtro2 = st.columns(2)
    
    with col_filtro1:
        areas_disponibles = ["Todas las áreas"] + sorted(df["AREA"].unique().tolist())
        area_seleccionada = st.selectbox("Seleccionar área:", areas_disponibles)
    
    with col_filtro2:
        meses_disponibles = ["Todos los meses"] + sorted(df["MES_NOMBRE"].unique().tolist())
        mes_seleccionado = st.selectbox("Seleccionar mes:", meses_disponibles)
    
    # Aplicar filtros
    df_filtrado = df.copy()
    if area_seleccionada != "Todas las áreas":
        df_filtrado = df_filtrado[df_filtrado["AREA"] == area_seleccionada]
    if mes_seleccionado != "Todos los meses":
        df_filtrado = df_filtrado[df_filtrado["MES_NOMBRE"] == mes_seleccionado]

    st.subheader("📊 Resultados del cálculo")
    
    # Columnas a mostrar con desglose detallado de los 3 tipos
    columnas_mostrar = [
        "FECHA", "DÍA", "NOMBRE", "CARGO", "CÉDULA", "AREA", 
        "TURNO", "TURNO ENTRADA", "TURNO SALIDA",           # Horarios del turno
        "HRA INGRESO", "HORA SALIDA", "HORAS TRABAJADAS_DISPLAY",
        "HORAS EXTRA DIURNA_DISPLAY", "VALOR EXTRA DIURNA",         # Extra diurna
        "HORAS EXTRA NOCTURNA_DISPLAY", "VALOR EXTRA NOCTURNA",     # Extra nocturna  
        "RECARGO NOCTURNO_DISPLAY", "VALOR RECARGO NOCTURNO",       # Recargo nocturno
        "TOTAL HORAS EXTRA_DISPLAY",                                 # Total de horas extras
        "ACTIVIDAD DESARROLLADA",
        "VALOR TOTAL EXTRAS",                                        # Total valor extras
        "COMISIÓN O BONIFICACIÓN"
    ]
    
    # Filtrar solo las columnas que existen
    columnas_disponibles = [col for col in columnas_mostrar if col in df_filtrado.columns]
    
    # Crear copia para formatear la visualización
    df_display = df_filtrado[columnas_disponibles].copy()
    
    # Renombrar columnas para mayor claridad (quitar _DISPLAY del nombre)
    renombrar = {
        "HORAS TRABAJADAS_DISPLAY": "HORAS TRABAJADAS",
        "HORAS EXTRA DIURNA_DISPLAY": "HORAS EXTRA DIURNA",
        "HORAS EXTRA NOCTURNA_DISPLAY": "HORAS EXTRA NOCTURNA",
        "RECARGO NOCTURNO_DISPLAY": "RECARGO NOCTURNO",
        "TOTAL HORAS EXTRA_DISPLAY": "TOTAL HORAS EXTRA"
    }
    df_display = df_display.rename(columns=renombrar)
    
    # Formatear valores monetarios para mejor visualización
    columnas_dinero = ["VALOR EXTRA DIURNA", "VALOR EXTRA NOCTURNA", "VALOR RECARGO NOCTURNO", 
                       "VALOR TOTAL EXTRAS", "COMISIÓN O BONIFICACIÓN"]
    
    for col in columnas_dinero:
        if col in df_display.columns:
            df_display[col] = df_display[col].apply(lambda x: f"${x:,.2f}" if pd.notna(x) else "$0.00")
    
    # Mostrar tabla con formato mejorado
    st.dataframe(df_display, use_container_width=True)

    # Visualizaciones
    col_viz1, col_viz2 = st.columns(2)
    
    with col_viz1:
        st.subheader("👤 Horas extra por empleado")
        fig1, ax1 = plt.subplots(figsize=(10, 6))
        empleado_stats = df_filtrado.groupby("NOMBRE").agg({
            "HORAS EXTRA DIURNA_DISPLAY": "sum",
            "HORAS EXTRA NOCTURNA_DISPLAY": "sum",
            "RECARGO NOCTURNO_DISPLAY": "sum"
        })
        if not empleado_stats.empty:
            empleado_stats.plot(kind='bar', ax=ax1, color=['#f37021', '#ff6600', '#ff9966'], stacked=True)
            ax1.set_ylabel("Horas")
            ax1.set_xlabel("Empleado")
            ax1.legend(["Extra Diurna", "Extra Nocturna", "Recargo Nocturno"])
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            st.pyplot(fig1)
            
            # Botón para descargar gráfico
            buf1 = BytesIO()
            fig1.savefig(buf1, format='png', dpi=300, bbox_inches='tight')
            buf1.seek(0)
            st.download_button(
                label="📊 Descargar gráfico empleados",
                data=buf1,
                file_name=f"grafico_empleados_{date.today().isoformat()}.png",
                mime="image/png",
                key="download_empleados"
            )
        else:
            st.info("No hay datos para mostrar con los filtros seleccionados")

    with col_viz2:
        st.subheader("🏢 Horas extra por área")
        fig2, ax2 = plt.subplots(figsize=(10, 6))
        area_stats = df_filtrado.groupby("AREA").agg({
            "HORAS EXTRA DIURNA_DISPLAY": "sum",
            "HORAS EXTRA NOCTURNA_DISPLAY": "sum",
            "RECARGO NOCTURNO_DISPLAY": "sum"
        })
        if not area_stats.empty:
            area_stats.plot(kind='barh', ax=ax2, color=['#f37021', '#ff6600', '#ff9966'], stacked=True)
            ax2.set_xlabel("Horas")
            ax2.set_ylabel("Área")
            ax2.legend(["Extra Diurna", "Extra Nocturna", "Recargo Nocturno"])
            plt.tight_layout()
            st.pyplot(fig2)
            
            # Botón para descargar gráfico
            buf2 = BytesIO()
            fig2.savefig(buf2, format='png', dpi=300, bbox_inches='tight')
            buf2.seek(0)
            st.download_button(
                label="📊 Descargar gráfico áreas",
                data=buf2,
                file_name=f"grafico_areas_{date.today().isoformat()}.png",
                mime="image/png",
                key="download_areas"
            )
        else:
            st.info("No hay datos para mostrar con los filtros seleccionados")

    # COMPARATIVO MENSUAL
    st.subheader("📅 Comparativo mensual")
    
    # Agrupar por mes
    comparativo_mensual = df.groupby("MES_NOMBRE").agg({
        "HORAS EXTRA DIURNA_DISPLAY": "sum",
        "HORAS EXTRA NOCTURNA_DISPLAY": "sum",
        "RECARGO NOCTURNO_DISPLAY": "sum",
        "VALOR TOTAL EXTRAS": "sum"
    }).reset_index()
    
    # Ordenar por fecha
    # Crear columna temporal con fecha para ordenar
    def extraer_fecha(mes_nombre):
        # De "Febrero 2025" extraer año y mes
        partes = mes_nombre.split()
        if len(partes) == 2:
            mes_texto, anio = partes
            meses = {
                'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
                'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
                'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
            }
            mes_num = meses.get(mes_texto, 1)
            return datetime(int(anio), mes_num, 1)
        return datetime(2025, 1, 1)
    
    comparativo_mensual["ORDEN"] = comparativo_mensual["MES_NOMBRE"].apply(extraer_fecha)
    comparativo_mensual = comparativo_mensual.sort_values("ORDEN")
    
    col_comp1, col_comp2 = st.columns(2)
    
    with col_comp1:
        st.markdown("#### Horas por mes")
        fig3, ax3 = plt.subplots(figsize=(10, 6))
        x = range(len(comparativo_mensual))
        width = 0.25
        
        ax3.bar([i - width for i in x], comparativo_mensual["HORAS EXTRA DIURNA_DISPLAY"], 
                width, label='Extra Diurna', color='#f37021')
        ax3.bar(x, comparativo_mensual["HORAS EXTRA NOCTURNA_DISPLAY"], 
                width, label='Extra Nocturna', color='#ff6600')
        ax3.bar([i + width for i in x], comparativo_mensual["RECARGO NOCTURNO_DISPLAY"], 
                width, label='Recargo Nocturno', color='#ff9966')
        
        ax3.set_xlabel('Mes')
        ax3.set_ylabel('Horas')
        ax3.set_xticks(x)
        ax3.set_xticklabels(comparativo_mensual["MES_NOMBRE"], rotation=45, ha='right')
        ax3.legend()
        plt.tight_layout()
        st.pyplot(fig3)
        
        # Botón para descargar gráfico de horas
        buf3 = BytesIO()
        fig3.savefig(buf3, format='png', dpi=300, bbox_inches='tight')
        buf3.seek(0)
        st.download_button(
            label="📊 Descargar gráfico de horas",
            data=buf3,
            file_name=f"grafico_horas_mensual_{date.today().isoformat()}.png",
            mime="image/png"
        )
    
    with col_comp2:
        st.markdown("#### Costos por mes")
        fig4, ax4 = plt.subplots(figsize=(10, 6))
        
        # Gráfico de barras simple
        x = range(len(comparativo_mensual))
        
        ax4.bar(x, comparativo_mensual["VALOR TOTAL EXTRAS"], 
                color='#f37021', label='Valor Total Extras')
        
        ax4.set_xlabel('Mes')
        ax4.set_ylabel('Valor ($)')
        ax4.set_xticks(x)
        ax4.set_xticklabels(comparativo_mensual["MES_NOMBRE"], rotation=45, ha='right')
        ax4.legend()
        plt.grid(True, alpha=0.3, axis='y')
        plt.tight_layout()
        st.pyplot(fig4)
        
        # Botón para descargar gráfico de costos
        buf4 = BytesIO()
        fig4.savefig(buf4, format='png', dpi=300, bbox_inches='tight')
        buf4.seek(0)
        st.download_button(
            label="📊 Descargar gráfico de costos",
            data=buf4,
            file_name=f"grafico_costos_mensual_{date.today().isoformat()}.png",
            mime="image/png"
        )
    
    # Tabla comparativa mensual
    st.markdown("#### Tabla comparativa mensual")
    tabla_comparativa = comparativo_mensual[["MES_NOMBRE", "HORAS EXTRA DIURNA_DISPLAY", "HORAS EXTRA NOCTURNA_DISPLAY",
                                             "RECARGO NOCTURNO_DISPLAY", "VALOR TOTAL EXTRAS"]].copy()
    tabla_comparativa.columns = ["Mes", "H. Extra Diurna", "H. Extra Nocturna", "H. Recargo Nocturno",
                                  "Valor Total Extras ($)"]
    
    # Formatear valores monetarios
    tabla_comparativa["Valor Total Extras ($)"] = tabla_comparativa["Valor Total Extras ($)"].apply(lambda x: f"${x:,.2f}")
    tabla_comparativa["H. Extra Diurna"] = tabla_comparativa["H. Extra Diurna"].apply(lambda x: f"{x:.2f}")
    tabla_comparativa["H. Extra Nocturna"] = tabla_comparativa["H. Extra Nocturna"].apply(lambda x: f"{x:.2f}")
    tabla_comparativa["H. Recargo Nocturno"] = tabla_comparativa["H. Recargo Nocturno"].apply(lambda x: f"{x:.2f}")
    
    st.dataframe(tabla_comparativa, use_container_width=True, hide_index=True)

    # Resumen estadístico (con datos filtrados)
    st.subheader("📈 Resumen general")
    col_stat1, col_stat2, col_stat3, col_stat4, col_stat5, col_stat6 = st.columns(6)
    
    with col_stat1:
        st.metric("H. Extra Diurnas", f"{df_filtrado['HORAS EXTRA DIURNA_DISPLAY'].sum():.2f}")
    
    with col_stat2:
        st.metric("H. Extra Nocturnas", f"{df_filtrado['HORAS EXTRA NOCTURNA_DISPLAY'].sum():.2f}")
    
    with col_stat3:
        st.metric("H. Recargo Nocturno", f"{df_filtrado['RECARGO NOCTURNO_DISPLAY'].sum():.2f}")
    
    with col_stat4:
        st.metric("Total Extras ($)", f"${df_filtrado['VALOR TOTAL EXTRAS'].sum():,.2f}")
    
    with col_stat5:
        st.metric("Total Bonificaciones ($)", f"${df_filtrado['COMISIÓN O BONIFICACIÓN'].sum():,.2f}")
    
    with col_stat6:
        st.metric("Total General ($)", f"${df_filtrado['VALOR TOTAL EXTRAS'].sum():,.2f}")

    # Descarga de resultados
    st.subheader("💾 Descargar resultados")
    
    col_desc1, col_desc2 = st.columns(2)
    
    with col_desc1:
        st.markdown("### 📥 Descargar datos completos en Excel")
        st.markdown("Incluye todos los cálculos y resultados detallados")
        
        # Crear el archivo Excel en memoria usando pandas directamente
        hoy = date.today().isoformat()
        output_filename = f"resultado_pagos_{hoy}.xlsx"
        
        # Preparar DataFrame para exportar con columnas en el orden exacto solicitado
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"
        
        # Definir encabezados en el orden EXACTO solicitado
        headers = [
            "CÉDULA", "NOMBRE", "CARGO", "AREA", "SALARIO BASICO", "COMISIÓN O BONIFICACIÓN",
            "TOTAL BASE LIQUIDACION", "Valor Ordinario Hora", "FECHA", "DÍA", "TURNO",
            "TURNO ENTRADA", "TURNO SALIDA", "HORA REAL INGRESO", "HORA REAL SALIDA",
            "ACTIVIDAD DESARROLLADA", "HORAS TRABAJADAS",
            "Cant. HORAS EXTRA DIURNA", "VALOR EXTRA DIURNA",
            "Cant. HORAS EXTRA NOCTURNA", "VALOR EXTRA NOCTURNA",
            "Cant. RECARGO NOCTURNO", "VALOR RECARGO NOCTURNO",
            "TOTAL HORAS EXTRA", "VALOR TOTAL EXTRAS",
            "MES", "MES_NOMBRE", "Observacion"
        ]
        
        ws.append(headers)
        
        # Formatear encabezados - BLANCO Y NEGRO (sin colores)
        header_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Mapeo de columnas del DataFrame a los encabezados
        column_mapping = {
            "CÉDULA": "CÉDULA",
            "NOMBRE": "NOMBRE",
            "CARGO": "CARGO",
            "AREA": "AREA",
            "SALARIO BASICO": "SALARIO BASICO",
            "COMISIÓN O BONIFICACIÓN": "COMISIÓN O BONIFICACIÓN",
            "TOTAL BASE LIQUIDACION": "TOTAL BASE LIQUIDACION",  # Calculado arriba
            "Valor Ordinario Hora": "IMPORTE HORA",
            "FECHA": "FECHA",
            "DÍA": "DÍA",
            "TURNO": "TURNO",
            "TURNO ENTRADA": "TURNO ENTRADA",
            "TURNO SALIDA": "TURNO SALIDA",
            "HORA REAL INGRESO": "HRA INGRESO",  # Columna real del DataFrame
            "HORA REAL SALIDA": "HORA SALIDA",    # Columna real del DataFrame
            "ACTIVIDAD DESARROLLADA": "ACTIVIDAD DESARROLLADA",
            "HORAS TRABAJADAS": "HORAS TRABAJADAS",
            "Cant. HORAS EXTRA DIURNA": "HORAS EXTRA DIURNA",
            "VALOR EXTRA DIURNA": "VALOR EXTRA DIURNA",
            "Cant. HORAS EXTRA NOCTURNA": "HORAS EXTRA NOCTURNA",
            "VALOR EXTRA NOCTURNA": "VALOR EXTRA NOCTURNA",
            "Cant. RECARGO NOCTURNO": "RECARGO NOCTURNO",
            "VALOR RECARGO NOCTURNO": "VALOR RECARGO NOCTURNO",
            "TOTAL HORAS EXTRA": "TOTAL HORAS EXTRA",
            "VALOR TOTAL EXTRAS": "VALOR TOTAL EXTRAS",
            "MES": "MES",
            "MES_NOMBRE": "MES_NOMBRE",
            "Observacion": "OBSERVACIONES"  # Nueva columna
        }
        
        # Escribir datos fila por fila
        for idx, row in df.iterrows():
            row_data = []
            for header in headers:
                df_col = column_mapping.get(header, header)
                
                if df_col in df.columns:
                    value = row[df_col]
                    
                    # Convertir valores según tipo a tipos básicos de Python
                    if pd.isna(value):
                        row_data.append("")
                    elif isinstance(value, (time, datetime)):
                        if isinstance(value, datetime):
                            row_data.append(value.strftime('%Y-%m-%d %H:%M'))
                        else:
                            row_data.append(value.strftime('%H:%M'))
                    elif isinstance(value, (np.int64, np.int32, np.int16, np.int8)):
                        row_data.append(int(value))
                    elif isinstance(value, (np.float64, np.float32, np.float16)):
                        # IMPORTANTE: NO redondear aquí, mantener precisión completa
                        # El formato de celda en Excel se encargará de mostrar 2 decimales
                        row_data.append(float(value))
                    elif isinstance(value, (pd.Period,)):
                        row_data.append(str(value))
                    elif hasattr(value, 'item'):  # numpy types
                        row_data.append(value.item())
                    else:
                        # Convertir a string si es cualquier otro tipo
                        try:
                            # Intentar convertir a tipo básico
                            if isinstance(value, (int, float, str, bool)):
                                row_data.append(value)
                            else:
                                row_data.append(str(value))
                        except:
                            row_data.append(str(value))
                else:
                    # Columna no existe, poner vacío
                    row_data.append("")
            
            ws.append(row_data)
        
        # Aplicar formato a todas las celdas - OPTIMIZADO
        # Pre-calcular el formato de cada columna UNA sola vez
        center_alignment = Alignment(horizontal="center", vertical="center")
        col_formats = {}
        for col_idx, header_name in enumerate(headers, 1):
            if "VALOR" in header_name or "SALARIO" in header_name or "TOTAL BASE" in header_name or "Valor Ordinario" in header_name:
                col_formats[col_idx] = '#,##0.00'
            elif "Cant." in header_name or "HORAS" in header_name:
                col_formats[col_idx] = '0.00'
            else:
                col_formats[col_idx] = None
        
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = center_alignment
                
                fmt = col_formats[col_idx]
                if fmt and isinstance(cell.value, (int, float)):
                    cell.number_format = fmt
        
        # Ajustar anchos de columnas
        column_widths = {
            'A': 12,  # CÉDULA
            'B': 25,  # NOMBRE
            'C': 20,  # CARGO
            'D': 15,  # AREA
            'E': 15,  # SALARIO BASICO
            'F': 18,  # COMISIÓN
            'G': 18,  # TOTAL BASE
            'H': 15,  # Valor Hora
            'I': 12,  # FECHA
            'J': 12,  # DÍA
            'K': 22,  # TURNO
            'L': 12,  # TURNO ENTRADA
            'M': 12,  # TURNO SALIDA
            'N': 14,  # HORA REAL INGRESO
            'O': 14,  # HORA REAL SALIDA
            'P': 30,  # ACTIVIDAD
            'Q': 12,  # HORAS TRAB
            'R': 12,  # Cant Extra Diurna
            'S': 15,  # Valor Extra Diurna
            'T': 12,  # Cant Extra Nocturna
            'U': 15,  # Valor Extra Nocturna
            'V': 12,  # Cant Recargo
            'W': 15,  # Valor Recargo
            'X': 12,  # Total Horas Extra
            'Y': 18,  # Valor Total Extras
            'Z': 10,  # MES
            'AA': 15,  # MES_NOMBRE
            'AB': 30,  # Observacion
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Guardar en buffer
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        st.download_button(
            label="📥 DESCARGAR EXCEL (.XLSX) ⬇️",
            data=excel_buffer,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel",
            use_container_width=True
        )
        
        st.warning("""
        ⚠️ **IMPORTANTE:** Si el archivo se descarga como `.csv`:
        
        **Solución rápida:**
        1. Descarga el archivo (aunque sea .csv)
        2. Renómbralo: cambia `.csv` por `.xlsx`
        3. Ábrelo en Excel - funcionará perfectamente
        
        El contenido **SÍ es Excel válido**, solo la extensión está incorrecta.
        """)
        
        st.info("""
        💡 **¿Por qué pasa esto?**
        
        Es un problema conocido de algunos navegadores (Chrome/Edge en Windows) con Streamlit.
        El archivo generado es 100% Excel, pero el navegador le pone extensión .csv al descargar.
        """)
    
    with col_desc2:
        st.markdown("### 📊 Todas las gráficas ya tienen su botón")
        st.markdown("""
        Cada gráfico tiene su propio botón de descarga:
        - Gráfico de empleados
        - Gráfico de áreas  
        - Gráfico de horas mensuales
        - Gráfico de costos mensuales
        """)
